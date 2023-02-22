from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
from database.models import RemoteSite
from netmiko import ConnectHandler
from netmiko import  NetMikoTimeoutException
from paramiko.ssh_exception import SSHException 
from paramiko.ssh_exception import  AuthenticationException
import openpyxl
from openpyxl.styles import PatternFill
import time
import re
from datetime import datetime
import random

#======================================================================================================



def ConnectSSH(request):
    if request.user.is_authenticated:
        wb = openpyxl.load_workbook("static\demo.xlsx")
        sheet = wb.active
        commandFile = open("static\\VerificationCommands.txt","r")
        commandList = commandFile.readlines()

        for i in range(2,sheet.max_row+1):
            Start_At = datetime.now().strftime("%H:%M:%S")
            data = sheet.cell(i,1).value
            print ("DATA:- ",data)
            if data == "#N/A":
                # mark the error and skip this branch
                sheet.cell(i,2).value = "IP Address Not Found"
                sheet.cell(i,2).fill=PatternFill("solid","0000ff")
                print("IP address issue : Marked in Excel")
                continue;
            if data is not None:
                if len(re.findall("/",data)) > 0:
                    ip = data[:len(data)-3]
                    #this will retrieve only the ip excluding /24 or /23 ending charachters
                else:
                    ip = data
            else:
                print("Sheet End OR No Data found")
                continue
            # data in correct format

            print(ip)

            try:
                net_connect = ConnectHandler(
                    device_type="cisco_xe",
                    host=ip,
                    username="nocsupport",
                    password="Goldsberg@!@#$%",
                    secret = "Sunr!se&*(",
                )
                # time.sleep(10)
                # hostname = net_connect.find_prompt()

            except (AuthenticationException):
                print ('Authentication Failure: {}'.format(ip))
                sheet.cell(i,2).value = "Authentication Failure"
                sheet.cell(i,2).fill=PatternFill("solid","FF0000")
                print("\n Please note:- this device has not been configured. \n")
                continue 
            except (NetMikoTimeoutException):
                print ('\n' + 'Timeout to device: {}'.format(ip))
                sheet.cell(i,2).value = "Timeout to device"
                sheet.cell(i,2).fill=PatternFill("solid","FF0000")
                print("\n Please note:- this device has not been configured. \n")
                continue
            except (SSHException):
                print ('SSH might not be enabled: {}'.format(ip))
                sheet.cell(i,2).value = "SSH might not be enabled"
                sheet.cell(i,2).fill=PatternFill("solid","FF0000")
                print("\n Please note:- this device has not been configured. \n")
                continue 
            except (EOFError):
                print ('\n' + 'End of attempting device: {}'.format(ip))
                sheet.cell(i,2).value = "End of attempting device"
                sheet.cell(i,2).fill=PatternFill("solid","FF0000")
                print("\n Please note:- this device has not been configured. \n")
                continue
            except Exception as unknown_error:
                print ("Some other error:{} for {}".format(str(unknown_error), ip)) 
                sheet.cell(i,2).value = "something Wrong with this site"
                sheet.cell(i,2).fill=PatternFill("solid","FF0000")
                print("\n Please note:- this device has not been configured. \n")
                continue

            print("ssh conncetion established for ",ip)

            try:
                out = open("static\\output\\{}-output.txt".format(str(ip)),'w')
                for line in commandList:
                    output = net_connect.send_command("{}".format(line))
                    # time.sleep(0.3)
                    out.writelines(net_connect.find_prompt())
                    out.write("\n")
                    out.write(line)
                    out.write("\n")
                    out.writelines(output)
                    print(line)
                out.close()

                out = open("static\\output\\{}-output.txt".format(str(ip)),'r')
                linesForTest = out.readlines()
                for line in linesForTest:
                    if "CPU utilization" in line:
                        numbers = re.findall(r'\d+\.\d+|\d+', line)
                        numbers.pop(1)
                        rownum = 3
                        for num in numbers:
                            if int(num) > 8:
                                print("high {}%".format(num))
                                sheet.cell(i,rownum).value = num
                                sheet.cell(i,rownum).fill=PatternFill("solid","FF0000")

                            else:
                                print("low {}%".format(num))
                                sheet.cell(i,rownum).value = num
                                sheet.cell(i,rownum).fill=PatternFill("solid","00FF00")
                            rownum += 1
                        break
                                
                    
                out.close()
                sheet.cell(i,2).value = "Retrieved"
                # print("Configurations has been saved successfull on {}".format(hostname))
            except Exception as e:
                sheet.cell(i,2).value = "something Wrong with this site"
                sheet.cell(i,2).fill=PatternFill("solid","FF0000")
                print("Router not configured : ",e)
            
        wb.save("static\WithRemark.xlsx")
        df = pd.read_excel('static\WithRemark.xlsx')
        headers = list(df.columns)
        data = df.values.tolist()
        return render(request, 'excel.html', {'headers': headers, 'data': data})
    else:
        messages.warning(request,"Login First")
        return redirect("/")



#=========================================================================================================

def TestExcel(request):
    if request.user.is_authenticated:
        wb = openpyxl.load_workbook("static/datasheet.xlsx")
        sheet = wb.active
        
        for i in range(2,sheet.max_row+1):
            site = RemoteSite(
                BranchName = sheet.cell(i,3).value,
                BranchCode = sheet.cell(i,4).value,
                ip = sheet.cell(i,11).value,
                lastFiveSecCpu = random.randint(40,100),
                lastOneMinCpu = random.randint(40,100),
                lastFiveMinCpu = random.randint(40,100),
                primaryPower = "ok",
                secondaryPower = "ok",
                fan = "ok",
                temp = random.randint(10,50)
            )
            site.save()
        
        # messages.success(request,"records inserted successfully")
        return redirect("/")
    else:
        messages.warning(request,"Login First")
        return redirect("/")



def SODEOD(request):

    if request.user.is_authenticated:
        siteData = RemoteSite.objects.all().order_by("-lastFiveSecCpu","-lastOneMinCpu","-lastFiveMinCpu")
        
        return render(request,'SODEOD.html',{"siteData":siteData})
    else:
        messages.warning(request,"Login First..")
        return redirect("/")



def UserLogin(request):
    if request.method == "POST":
        username = request.POST.get('Username')
        password = request.POST.get('Password')
        try:
            user = authenticate(username=username, password=password) # authenticating user
            login(request, user) # login user if authenticated
            return redirect('/')
        except Exception as e:
            messages.error(request, "Authentication Faild..")  # return error if user not authenticated
            return redirect('/')
    elif request.user.is_authenticated:
        messages.info(request,"You are already logged in")
        return redirect("/")
    else:
        notice = open("static\\notice.txt",'r').readlines()
        return render(request,'login.html',{"notice":notice})

def logoutUser(request):
    logout(request)
    return redirect("/")








def index(request):
    if request.user.is_authenticated:
        return render(request,"index.html",{"active_Home":"active"})
    else:
        return redirect("/login")